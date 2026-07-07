"""
GPU Insight Lab - Excel Report Generator
Generates management-readable Excel workbook using openpyxl.
Sheets: Summary, System, Benchmarks, Diagnostics, Raw Measurements, Comparison, Toolchain.
"""

from __future__ import annotations

import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

from app.branding import APP_NAME, APP_VERSION, REPORT_PREFIX

logger = logging.getLogger(__name__)


def _safe_val(v: Any, decimals: int = 3) -> Any:
    """Return clean value for Excel cell."""
    if v is None:
        return ""
    if isinstance(v, float):
        return round(v, decimals)
    if isinstance(v, bool):
        return "Yes" if v else "No"
    if isinstance(v, list):
        return str(v)
    return v


def generate(
    session: Any,
    output_dir: Optional[Path] = None,
) -> Path:
    """Generate Excel report. Returns path to created file."""
    try:
        import openpyxl  # type: ignore
        from openpyxl.styles import (  # type: ignore
            Alignment,
            Font,
            PatternFill,
            Border,
            Side,
        )
        from openpyxl.utils import get_column_letter  # type: ignore
    except ImportError as exc:
        logger.error("openpyxl not installed; cannot generate Excel report: %s", exc)
        raise

    if output_dir is None:
        from app.config import get_config  # noqa: PLC0415
        output_dir = get_config().output_path()

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    ts = datetime.now(tz=timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"{REPORT_PREFIX}_{ts}.xlsx"
    file_path = output_dir / filename

    # Extract data
    if hasattr(session, "to_dict"):
        data = session.to_dict()
    elif isinstance(session, dict):
        data = session
    else:
        data = {}

    sys_info = data.get("system_info", {}) or {}
    nv_info = data.get("nvidia_info", {}) or {}
    pcie_info = data.get("pcie_info", {}) or {}
    cuda_info = data.get("cuda_info", {}) or {}
    tools = data.get("tool_status", {}) or {}
    results_raw = data.get("results", []) or []
    diag_results = data.get("diagnosis_results", []) or []
    score = data.get("health_score")
    confidence = data.get("score_confidence")
    score_details = data.get("score_details", {}) or {}

    results: List[Dict] = []
    for r in results_raw:
        if isinstance(r, dict):
            results.append(r)
        elif hasattr(r, "to_dict"):
            results.append(r.to_dict())

    # Style helpers
    HEADER_FILL = PatternFill("solid", fgColor="1A237E")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    TITLE_FONT = Font(size=14, bold=True, color="1A237E")
    SUB_FONT = Font(size=11, bold=True)
    BORDER = Border(
        bottom=Side(style="thin"),
        right=Side(style="thin"),
    )

    SEV_FILL = {
        "INFO": PatternFill("solid", fgColor="E1F5FE"),
        "WARNING": PatternFill("solid", fgColor="FFF8E1"),
        "ERROR": PatternFill("solid", fgColor="FFEBEE"),
        "CRITICAL": PatternFill("solid", fgColor="FFEBEE"),
    }
    SEV_FONT = {
        "INFO": Font(color="0277BD"),
        "WARNING": Font(color="F57F17"),
        "ERROR": Font(color="C62828"),
        "CRITICAL": Font(color="B71C1C", bold=True),
    }

    def _style_header_row(ws: Any, row: int, num_cols: int) -> None:
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _auto_width(ws: Any) -> None:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:  # noqa: BLE001
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 60)

    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ SUMMARY
    ws_sum = wb.active
    ws_sum.title = "Summary"

    ws_sum["A1"] = "GPU Insight Lab Benchmark and Diagnostic Report"
    ws_sum["A1"].font = TITLE_FONT
    ws_sum["A2"] = f"Generated: {datetime.now(tz=timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}"
    ws_sum["A3"] = f"Tool: {APP_NAME} v{APP_VERSION}"
    ws_sum["A4"] = f"Session: {data.get('session_name', '')} | Status: {data.get('status', '')}"
    ws_sum.merge_cells("A1:E1")

    row = 6
    ws_sum.cell(row, 1, "GPU Insight Score").font = SUB_FONT
    ws_sum.cell(row + 1, 1, "Score (0-100)")
    ws_sum.cell(row + 1, 2, _safe_val(score, 1))
    ws_sum.cell(row + 2, 1, "Confidence")
    ws_sum.cell(row + 2, 2, f"{(confidence or 0)*100:.0f}%" if confidence is not None else "")

    row = 10
    ws_sum.cell(row, 1, "Category Scores").font = SUB_FONT
    cat_scores = score_details.get("category_scores", {}) or {}
    for cat, pts in cat_scores.items():
        row += 1
        ws_sum.cell(row, 1, cat.replace("_", " ").title())
        ws_sum.cell(row, 2, _safe_val(pts, 1))

    row += 2
    ws_sum.cell(row, 1, "Positive Findings").font = SUB_FONT
    for pf in score_details.get("positive_findings", []) or []:
        row += 1
        ws_sum.cell(row, 1, f"  {pf}")

    row += 2
    ws_sum.cell(row, 1, "Missing Data").font = SUB_FONT
    for md in score_details.get("missing_data", []) or []:
        row += 1
        ws_sum.cell(row, 1, f"  {md}")

    ws_sum.freeze_panes = "A2"
    _auto_width(ws_sum)

    # ------------------------------------------------------------------ SYSTEM
    ws_sys = wb.create_sheet("System")
    headers = ["Field", "Value"]
    for ci, h in enumerate(headers, 1):
        ws_sys.cell(1, ci, h)
    _style_header_row(ws_sys, 1, len(headers))
    ws_sys.auto_filter.ref = f"A1:B1"
    ws_sys.freeze_panes = "A2"

    system_fields = [
        ("Hostname", sys_info.get("hostname", "")),
        ("OS Name", sys_info.get("os_name", "")),
        ("OS Version", sys_info.get("os_version", "")),
        ("Architecture", sys_info.get("architecture", "")),
        ("CPU Model", sys_info.get("cpu_model", "")),
        ("CPU Physical Cores", sys_info.get("cpu_physical_cores", "")),
        ("CPU Logical Cores", sys_info.get("cpu_logical_cores", "")),
        ("RAM Total (GB)", _safe_val(sys_info.get("ram_total_gb"), 2)),
        ("RAM Available (GB)", _safe_val(sys_info.get("ram_available_gb"), 2)),
        ("Python Version", sys_info.get("python_version", "")),
        ("GPU Name", nv_info.get("gpu_name", "")),
        ("GPU UUID", nv_info.get("gpu_uuid", "")),
        ("Driver Version", nv_info.get("driver_version", "")),
        ("CUDA Driver Version", nv_info.get("cuda_driver_version", "")),
        ("Compute Capability", nv_info.get("compute_capability", "")),
        ("VRAM Total (MB)", _safe_val(nv_info.get("vram_total_mb"), 0)),
        ("VRAM Used (MB)", _safe_val(nv_info.get("vram_used_mb"), 0)),
        ("VRAM Free (MB)", _safe_val(nv_info.get("vram_free_mb"), 0)),
        ("Temperature (°C)", _safe_val(nv_info.get("temperature_c"), 1)),
        ("GPU Utilization (%)", _safe_val(nv_info.get("gpu_utilization_pct"), 1)),
        ("Power Draw (W)", _safe_val(nv_info.get("power_draw_w"), 1)),
        ("Power Limit (W)", _safe_val(nv_info.get("power_limit_w"), 1)),
        ("GPU Clock (MHz)", _safe_val(nv_info.get("gpu_clock_mhz"), 0)),
        ("Mem Clock (MHz)", _safe_val(nv_info.get("mem_clock_mhz"), 0)),
        ("Performance State", nv_info.get("performance_state", "")),
        ("PCIe Gen Current", _safe_val(pcie_info.get("pcie_gen_current"))),
        ("PCIe Width Current", _safe_val(pcie_info.get("pcie_width_current"))),
        ("PCIe Gen Max", _safe_val(pcie_info.get("pcie_gen_max"))),
        ("PCIe Width Max", _safe_val(pcie_info.get("pcie_width_max"))),
        ("PCIe BW Current (GB/s)", _safe_val(pcie_info.get("bandwidth_gbps_current"), 2)),
        ("PCIe BW Theoretical (GB/s)", _safe_val(pcie_info.get("bandwidth_gbps_theoretical"), 2)),
        ("NVCC Available", _safe_val(cuda_info.get("nvcc_available"))),
        ("NVCC Version", cuda_info.get("nvcc_version", "")),
        ("CUDA Runtime Available", _safe_val(cuda_info.get("cuda_runtime_available"))),
        ("CUDA Runtime Version", cuda_info.get("cuda_runtime_version", "")),
        ("Native Benchmark", _safe_val(cuda_info.get("native_benchmark_available"))),
    ]
    for row_i, (field, value) in enumerate(system_fields, 2):
        ws_sys.cell(row_i, 1, field)
        ws_sys.cell(row_i, 2, value)

    _auto_width(ws_sys)

    # ------------------------------------------------------------------ BENCHMARKS
    ws_bm = wb.create_sheet("Benchmarks")
    bm_headers = [
        "Test Name", "Data Type", "Input Size", "Block Size", "Warmup", "Measured",
        "CPU Time (ms)", "GPU Time (ms)", "Transfer Time (ms)", "End-to-End (ms)",
        "Throughput", "Bandwidth (GB/s)", "Speedup", "Correctness", "Max Error",
        "Mean (ms)", "Median (ms)", "Min (ms)", "Max (ms)", "Std Dev",
        "GPU Name", "Driver", "Notes", "Error",
    ]
    for ci, h in enumerate(bm_headers, 1):
        ws_bm.cell(1, ci, h)
    _style_header_row(ws_bm, 1, len(bm_headers))
    ws_bm.auto_filter.ref = f"A1:{get_column_letter(len(bm_headers))}1"
    ws_bm.freeze_panes = "A2"

    for ri, r in enumerate(results, 2):
        vals = [
            r.get("test_name", ""),
            r.get("data_type", ""),
            _safe_val(r.get("input_size")),
            _safe_val(r.get("block_size")),
            _safe_val(r.get("warmup_runs")),
            _safe_val(r.get("measured_runs")),
            _safe_val(r.get("cpu_time_ms")),
            _safe_val(r.get("gpu_time_ms")),
            _safe_val(r.get("transfer_time_ms")),
            _safe_val(r.get("end_to_end_time_ms")),
            _safe_val(r.get("throughput")),
            _safe_val(r.get("bandwidth_gbps")),
            _safe_val(r.get("speedup")),
            "Pass" if r.get("correctness_pass") is True else ("Fail" if r.get("correctness_pass") is False else "N/A"),
            _safe_val(r.get("max_error")),
            _safe_val(r.get("mean")),
            _safe_val(r.get("median")),
            _safe_val(r.get("min_val")),
            _safe_val(r.get("max_val")),
            _safe_val(r.get("standard_deviation")),
            r.get("gpu_name", ""),
            r.get("driver_version", ""),
            r.get("notes", ""),
            r.get("error", ""),
        ]
        for ci, v in enumerate(vals, 1):
            ws_bm.cell(ri, ci, v)

    _auto_width(ws_bm)

    # ------------------------------------------------------------------ DIAGNOSTICS
    ws_diag = wb.create_sheet("Diagnostics")
    diag_headers = [
        "Rule ID", "Severity", "Category", "Title", "Summary",
        "Evidence", "Confidence", "Recommendation", "Verification Step",
    ]
    for ci, h in enumerate(diag_headers, 1):
        ws_diag.cell(1, ci, h)
    _style_header_row(ws_diag, 1, len(diag_headers))
    ws_diag.auto_filter.ref = f"A1:{get_column_letter(len(diag_headers))}1"
    ws_diag.freeze_panes = "A2"

    for ri, d in enumerate(diag_results, 2):
        sev = d.get("severity", "INFO")
        vals = [
            d.get("rule_id", ""),
            sev,
            d.get("category", ""),
            d.get("title", ""),
            d.get("summary", ""),
            d.get("evidence", ""),
            _safe_val(d.get("confidence"), 2),
            d.get("recommendation", ""),
            d.get("verification_step", ""),
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws_diag.cell(ri, ci, v)
            if ci == 2:  # Severity column
                cell.fill = SEV_FILL.get(sev, PatternFill())
                cell.font = SEV_FONT.get(sev, Font())

    _auto_width(ws_diag)

    # ------------------------------------------------------------------ RAW MEASUREMENTS
    ws_raw = wb.create_sheet("Raw Measurements")
    ws_raw.cell(1, 1, "Test Name").font = HEADER_FONT
    ws_raw.cell(1, 1).fill = HEADER_FILL
    ws_raw.freeze_panes = "A2"

    row = 2
    for r in results:
        test_name = r.get("test_name", "")
        raw = r.get("raw_measurements", [])
        if not raw:
            continue
        ws_raw.cell(row, 1, test_name).font = SUB_FONT
        for i, v in enumerate(raw):
            ws_raw.cell(row, 2 + i, round(v, 4) if isinstance(v, float) else v)
        row += 1

    _auto_width(ws_raw)

    # ------------------------------------------------------------------ INTERVIEW DEMO SUMMARY
    ws_demo = wb.create_sheet("Interview Demo")
    ws_demo["A1"] = "GPU Insight Lab — Interview Demo Summary: Implemented vs. Roadmap"
    ws_demo["A1"].font = TITLE_FONT
    ws_demo.merge_cells("A1:D1")

    ws_demo["A3"] = "IMPLEMENTED FEATURES"
    ws_demo["A3"].font = Font(bold=True, color="2E7D32")
    implemented = [
        "System Inspector (CPU/GPU/PCIe/CUDA telemetry)",
        "Memory Benchmark (H2D/D2H/D2D via native CUDA binary)",
        "Kernel Lab (vector_add, reduction, transpose, gemm_naive, gemm_tiled, stream_pipeline)",
        "Evidence-based Diagnosis Engine (9 rules, evidence strings required)",
        "GPU Insight Score (0-100 composite across 6 categories)",
        "Multi-format Reports (JSON, CSV, Markdown, HTML, Excel)",
        "SQLite Session History with delta comparison",
        "PySide6 GUI (QMainWindow + QThread workers)",
        "CLI Automation (10 commands, --json flag)",
        "CUDA to HIP Portability Demo (vector_add_hip.cpp — NOT_VALIDATED on AMD hardware)",
    ]
    for i, feat in enumerate(implemented, 4):
        ws_demo.cell(i, 1, feat)

    roadmap_row = 4 + len(implemented) + 1
    ws_demo.cell(roadmap_row, 1, "ROADMAP / NOT YET IMPLEMENTED").font = Font(bold=True, color="C62828")
    roadmap = [
        "ROADMAP: softmax, layer_norm, GELU (AI inference kernels)",
        "ROADMAP: Flash Attention kernel",
        "ROADMAP: INT8 quantization",
        "ROADMAP: PyTorch extension integration",
        "ROADMAP: TensorRT plugin",
        "ROADMAP: cuFFT / cuBLAS full benchmark suite",
        "ROADMAP: Streamlit dashboard",
        "ROADMAP: Parquet storage backend",
        "ROADMAP: Multi-machine import",
        "ROADMAP: Company report templates",
        "ROADMAP: Batch execution / session manifests",
        "ROADMAP: Pass/fail policy YAML",
        "NOT_VALIDATED: AMD HIP real GPU benchmarks (requires ROCm hardware)",
    ]
    for i, feat in enumerate(roadmap, roadmap_row + 1):
        ws_demo.cell(i, 1, feat)

    _auto_width(ws_demo)

    # ------------------------------------------------------------------ COMPARISON (placeholder)
    ws_cmp = wb.create_sheet("Comparison")
    ws_cmp.cell(1, 1, "Comparison Sheet").font = TITLE_FONT
    ws_cmp.cell(2, 1, "Use the CLI 'compare' command to generate a side-by-side comparison report.")
    ws_cmp.cell(3, 1, "Select two sessions and re-generate this report in the Report Studio.")
    _auto_width(ws_cmp)

    # ------------------------------------------------------------------ TOOLCHAIN
    ws_tc = wb.create_sheet("Toolchain")
    tc_headers = ["Tool", "Available", "Version", "Path", "Error"]
    for ci, h in enumerate(tc_headers, 1):
        ws_tc.cell(1, ci, h)
    _style_header_row(ws_tc, 1, len(tc_headers))
    ws_tc.auto_filter.ref = f"A1:{get_column_letter(len(tc_headers))}1"
    ws_tc.freeze_panes = "A2"

    for ri, (tool_name, status) in enumerate(tools.items(), 2):
        if isinstance(status, dict):
            exists = status.get("exists", False)
            version = status.get("version", "")
            path = status.get("path", "")
            error = status.get("error", "")
        else:
            exists = False
            version = ""
            path = ""
            error = str(status)
        ws_tc.cell(ri, 1, tool_name)
        cell = ws_tc.cell(ri, 2, "Yes" if exists else "No")
        if exists:
            cell.font = Font(color="2E7D32")
        else:
            cell.font = Font(color="757575")
        ws_tc.cell(ri, 3, version)
        ws_tc.cell(ri, 4, path)
        ws_tc.cell(ri, 5, error)

    _auto_width(ws_tc)

    try:
        wb.save(str(file_path))
        logger.info("Excel report written to %s", file_path)
    except OSError as exc:
        logger.error("Failed to write Excel report: %s", exc)
        raise

    return file_path
