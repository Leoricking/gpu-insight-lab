"""
Tests for interview demo readiness — verifies reports contain expected content.
No GPU required.
"""

from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

ROOT = Path(__file__).parent.parent
SAMPLE_SESSION = ROOT / "examples" / "sample_session.json"


def _load_sample_session():
    with SAMPLE_SESSION.open(encoding="utf-8") as f:
        return json.load(f)


class TestReportTitle(unittest.TestCase):
    """Report title must include GPU Insight Lab."""

    def test_json_report_has_gpu_insight_lab(self) -> None:
        from app.branding import APP_NAME
        self.assertEqual(APP_NAME, "GPU Insight Lab")

    def test_markdown_title(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            from reports.markdown_report import generate
            session = _load_sample_session()
            p = generate(session, output_dir=Path(tmpdir))
            content = p.read_text(encoding="utf-8")
            self.assertIn("GPU Insight Lab", content)

    def test_html_report_title(self) -> None:
        try:
            import jinja2  # noqa: F401
        except ImportError:
            self.skipTest("jinja2 not installed")
        with tempfile.TemporaryDirectory() as tmpdir:
            from reports.html_report import generate
            session = _load_sample_session()
            p = generate(session, output_dir=Path(tmpdir))
            content = p.read_text(encoding="utf-8")
            self.assertIn("GPU Insight Lab", content)


class TestJSONReport(unittest.TestCase):
    """JSON report from sample data must contain required keys."""

    def test_json_report_keys(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            from reports.json_report import generate
            session = _load_sample_session()
            p = generate(session, output_dir=Path(tmpdir))
            data = json.loads(p.read_text(encoding="utf-8"))
            for key in ["session_id", "timestamp", "environment", "benchmarks", "diagnosis", "score"]:
                self.assertIn(key, data, f"JSON report missing key: {key}")

    def test_json_report_environment_keys(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            from reports.json_report import generate
            session = _load_sample_session()
            p = generate(session, output_dir=Path(tmpdir))
            data = json.loads(p.read_text(encoding="utf-8"))
            env = data.get("environment", {})
            self.assertIn("cuda_available", env)
            self.assertIn("gpu_name", env)


class TestMarkdownReport(unittest.TestCase):
    """Markdown report must contain GPU Insight Lab."""

    def test_contains_gpu_insight_lab(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            from reports.markdown_report import generate
            session = _load_sample_session()
            p = generate(session, output_dir=Path(tmpdir))
            self.assertIn("GPU Insight Lab", p.read_text(encoding="utf-8"))

    def test_contains_implemented_vs_roadmap(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            from reports.markdown_report import generate
            session = _load_sample_session()
            p = generate(session, output_dir=Path(tmpdir))
            content = p.read_text(encoding="utf-8")
            self.assertIn("Implemented vs. Roadmap", content)


class TestHTMLReport(unittest.TestCase):
    """HTML report must contain GPU Insight Lab."""

    def test_contains_gpu_insight_lab(self) -> None:
        try:
            import jinja2  # noqa: F401
        except ImportError:
            self.skipTest("jinja2 not installed")
        with tempfile.TemporaryDirectory() as tmpdir:
            from reports.html_report import generate
            session = _load_sample_session()
            p = generate(session, output_dir=Path(tmpdir))
            self.assertIn("GPU Insight Lab", p.read_text(encoding="utf-8"))


class TestExcelReport(unittest.TestCase):
    """Excel report must be openable by openpyxl."""

    def test_excel_opens_without_error(self) -> None:
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            self.skipTest("openpyxl not installed")
        with tempfile.TemporaryDirectory() as tmpdir:
            from reports.excel_report import generate
            session = _load_sample_session()
            p = generate(session, output_dir=Path(tmpdir))
            wb = openpyxl.load_workbook(str(p))
            self.assertIn("Summary", wb.sheetnames)

    def test_excel_has_interview_demo_sheet(self) -> None:
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            self.skipTest("openpyxl not installed")
        with tempfile.TemporaryDirectory() as tmpdir:
            from reports.excel_report import generate
            session = _load_sample_session()
            p = generate(session, output_dir=Path(tmpdir))
            wb = openpyxl.load_workbook(str(p))
            self.assertIn("Interview Demo", wb.sheetnames)


class TestDiagnosisFields(unittest.TestCase):
    """Diagnosis result must have required fields."""

    def test_diagnosis_result_fields(self) -> None:
        from diagnosis.engine import DiagnosisResult
        import dataclasses
        field_names = {f.name for f in dataclasses.fields(DiagnosisResult)}
        required = {
            "gpu_insight_score",
            "health_score",
            "confidence",
            "missing_data",
            "recommendations",
        }
        for field in required:
            self.assertIn(field, field_names,
                          f"DiagnosisResult missing field: {field}")

    def test_score_result_fields(self) -> None:
        from diagnosis.scoring import ScoreResult
        import dataclasses
        field_names = {f.name for f in dataclasses.fields(ScoreResult)}
        for field in ["score", "confidence", "missing_data", "positive_findings", "deductions"]:
            self.assertIn(field, field_names,
                          f"ScoreResult missing field: {field}")

    def test_sample_diagnosis_has_score(self) -> None:
        sample = ROOT / "examples" / "sample_diagnosis.json"
        with sample.open(encoding="utf-8") as f:
            data = json.load(f)
        self.assertIn("gpu_insight_score", data)
        self.assertIn("confidence", data)
        self.assertIn("missing_data", data)
        self.assertIn("recommendations", data)


if __name__ == "__main__":
    unittest.main()
