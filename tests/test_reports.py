"""
Tests for report generation with mock session data.
"""

import json
import tempfile
import unittest
from pathlib import Path


def _make_mock_session() -> dict:
    return {
        "session_id": "test-session-001",
        "session_name": "Test Session",
        "started_at": 1720000000.0,
        "completed_at": 1720000060.0,
        "status": "completed",
        "health_score": 72.5,
        "score_confidence": 0.80,
        "system_info": {
            "hostname": "test-machine",
            "os_name": "Windows",
            "os_release": "10",
            "cpu_model": "Intel Core i9-13900K",
            "cpu_logical_cores": 24,
            "ram_total_gb": 64.0,
            "python_version": "3.11.0",
        },
        "nvidia_info": {
            "available": True,
            "gpu_name": "NVIDIA Test GPU",
            "driver_version": "545.23",
            "cuda_driver_version": "12.3",
            "temperature_c": 65.0,
            "power_draw_w": 120.0,
            "vram_total_mb": 8192.0,
            "vram_used_mb": 2048.0,
            "gpu_utilization_pct": 85.0,
            "performance_state": "P0",
        },
        "cuda_info": {
            "nvcc_available": True,
            "nvcc_version": "12.3",
            "cuda_runtime_available": True,
            "native_benchmark_available": True,
        },
        "pcie_info": {
            "available": True,
            "pcie_gen_current": 4,
            "pcie_width_current": 16,
            "pcie_gen_max": 4,
            "pcie_width_max": 16,
            "bandwidth_gbps_current": 31.5,
            "bandwidth_gbps_theoretical": 31.5,
            "is_bottlenecked": False,
        },
        "tool_status": {
            "nvcc": {"exists": True, "version": "12.3", "path": "/usr/local/cuda/bin/nvcc"},
            "cmake": {"exists": True, "version": "3.27", "path": "/usr/bin/cmake"},
            "nvidia-smi": {"exists": True, "version": "545.23", "path": "/usr/bin/nvidia-smi"},
            "nsys": {"exists": False, "version": None, "path": None},
            "ncu": {"exists": False, "version": None, "path": None},
        },
        "results": [
            {
                "test_name": "cpu_vector_add",
                "mean": 1.234,
                "median": 1.220,
                "min_val": 1.100,
                "max_val": 1.500,
                "standard_deviation": 0.100,
                "bandwidth_gbps": 12.5,
                "correctness_pass": True,
                "raw_measurements": [1.1, 1.2, 1.3, 1.2, 1.1, 1.3, 1.2, 1.3, 1.5, 1.4],
                "notes": "CPU baseline",
                "gpu_name": "",
                "data_type": "float32",
                "input_size": 1000000,
            },
            {
                "test_name": "vector_add",
                "mean": 0.456,
                "median": 0.450,
                "min_val": 0.400,
                "max_val": 0.510,
                "standard_deviation": 0.030,
                "bandwidth_gbps": 95.3,
                "speedup": 2.7,
                "correctness_pass": True,
                "raw_measurements": [0.41, 0.45, 0.46, 0.45, 0.50, 0.46, 0.46, 0.46, 0.46, 0.46],
                "notes": "CUDA kernel",
                "gpu_name": "NVIDIA Test GPU",
                "data_type": "float32",
                "input_size": 16777216,
            },
        ],
        "diagnosis_results": [
            {
                "rule_id": "profiler_unavailable",
                "severity": "INFO",
                "category": "TOOLCHAIN_INCOMPLETE",
                "title": "Profiling Tools Not Available",
                "summary": "nsys and ncu are not installed.",
                "evidence": "Tool check: nsys=missing, ncu=missing",
                "confidence": 1.0,
                "recommendation": "Install Nsight Systems and Nsight Compute.",
                "verification_step": "Run nsys --version after installation.",
            },
        ],
        "score_details": {
            "score": 72.5,
            "confidence": 0.80,
            "missing_data": ["nsys", "ncu"],
            "positive_findings": ["NVIDIA GPU detected: NVIDIA Test GPU", "nvcc available"],
            "deductions": [{"category": "environment", "pts": -3, "reason": "nsys missing"}],
            "category_scores": {
                "environment_readiness": 16.0,
                "gpu_runtime_availability": 15.0,
                "kernel_correctness": 20.0,
            },
        },
    }


class TestJSONReport(unittest.TestCase):
    def test_generates_file(self) -> None:
        from reports.json_report import generate
        session = _make_mock_session()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate(session, output_dir=Path(tmpdir))
            self.assertTrue(path.exists())
            with path.open("r", encoding="utf-8") as fh:
                data = json.load(fh)
            self.assertIn("session", data)
            self.assertIn("report_metadata", data)

    def test_filename_format(self) -> None:
        from reports.json_report import generate
        session = _make_mock_session()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate(session, output_dir=Path(tmpdir))
            self.assertTrue(path.name.startswith("gpu_insight_session_"))
            self.assertTrue(path.name.endswith(".json"))


class TestCSVReport(unittest.TestCase):
    def test_generates_csv(self) -> None:
        from reports.csv_report import generate
        session = _make_mock_session()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate(session, output_dir=Path(tmpdir))
            self.assertTrue(path.exists())
            content = path.read_text(encoding="utf-8")
            self.assertIn("test_name", content)
            self.assertIn("cpu_vector_add", content)

    def test_has_header_row(self) -> None:
        from reports.csv_report import generate
        session = _make_mock_session()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate(session, output_dir=Path(tmpdir))
            lines = path.read_text(encoding="utf-8").splitlines()
            self.assertGreater(len(lines), 1)
            header = lines[0]
            self.assertIn("mean", header)
            self.assertIn("bandwidth_gbps", header)


class TestMarkdownReport(unittest.TestCase):
    def test_generates_markdown(self) -> None:
        from reports.markdown_report import generate
        session = _make_mock_session()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate(session, output_dir=Path(tmpdir))
            self.assertTrue(path.exists())
            content = path.read_text(encoding="utf-8")
            self.assertIn("# GPU Insight Lab Session Report", content)
            self.assertIn("## Benchmark Results", content)
            self.assertIn("## Diagnosis Results", content)

    def test_trademark_notice_present(self) -> None:
        from reports.markdown_report import generate
        session = _make_mock_session()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate(session, output_dir=Path(tmpdir))
            content = path.read_text(encoding="utf-8")
            self.assertIn("NVIDIA", content)


class TestHTMLReport(unittest.TestCase):
    def test_generates_html(self) -> None:
        try:
            import jinja2  # noqa: F401
        except ImportError:
            self.skipTest("Jinja2 not installed")
        from reports.html_report import generate
        session = _make_mock_session()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate(session, output_dir=Path(tmpdir))
            self.assertTrue(path.exists())
            content = path.read_text(encoding="utf-8")
            self.assertIn("<!DOCTYPE html>", content)
            self.assertIn("GPU Insight Lab", content)

    def test_html_has_score_section(self) -> None:
        try:
            import jinja2  # noqa: F401
        except ImportError:
            self.skipTest("Jinja2 not installed")
        from reports.html_report import generate
        session = _make_mock_session()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate(session, output_dir=Path(tmpdir))
            content = path.read_text(encoding="utf-8")
            self.assertIn("GPU Insight Score", content)
            self.assertIn("72", content)  # The score value


class TestExcelReport(unittest.TestCase):
    def test_generates_excel(self) -> None:
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            self.skipTest("openpyxl not installed")
        from reports.excel_report import generate
        session = _make_mock_session()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate(session, output_dir=Path(tmpdir))
            self.assertTrue(path.exists())
            self.assertTrue(path.name.endswith(".xlsx"))
            import openpyxl
            wb = openpyxl.load_workbook(str(path))
            self.assertIn("Summary", wb.sheetnames)
            self.assertIn("Benchmarks", wb.sheetnames)
            self.assertIn("Diagnostics", wb.sheetnames)
            self.assertIn("Toolchain", wb.sheetnames)


if __name__ == "__main__":
    unittest.main()
